from openpyxl import load_workbook, Workbook
from datetime import datetime
import sqlite3
import re
from .normalize import norm_inv

# Заголовки "как в исходном"
HEADERS = [
    "Наименование оборудования",
    "Дата принятия к учету",
    "Инвентарный номер",
    "Текущий владелец (сотрудник)",
    "Серийный номер",
    "Дата постановки / последняя дата актуализации (опционально)",
    "Текущий кабинет (кабинет)",
]


def _norm_header(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


# def import_reference_xlsx(conn: sqlite3.Connection, path: str) -> int:
#     wb = load_workbook(path)
#     ws = wb["os_reference"] if "os_reference" in wb.sheetnames else wb.active

#     print("IMPORT FILE:", path)
#     print("SHEET:", ws.title, "max_row:", ws.max_row, "max_col:", ws.max_column)

#     # ищем строку заголовков
#     header_row_idx = None
#     header_values = None

#     for r in range(1, 21):
#         row_vals = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
#         normed = [_norm_header(v) for v in row_vals]
#         if "инвентарный номер" in normed:
#             header_row_idx = r
#             header_values = row_vals
#             break

#     if header_row_idx is None:
#         raise RuntimeError("Не найдена строка заголовков (нет 'Инвентарный номер')")

#     # ✅ ОДИН раз, с правильным отступом
#     idx = {_norm_header(v): i for i, v in enumerate(header_values) if _norm_header(v)}

#     def get(row, col_name: str) -> str:
#         i = idx.get(_norm_header(col_name))
#         if i is None or i >= len(row):
#             return ""
#         v = row[i]
#         return "" if v is None else str(v).strip()

#     def get_any(row, *col_names: str) -> str:
#         for name in col_names:
#             v = get(row, name)
#             if v:
#                 return v
#         return ""

#     rows = []
#     for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
#         inv = norm_inv(get(row, "Инвентарный номер"))
#         if not inv:
#             continue

#         owner = get_any(row, "Текущий владелец (сотрудник)", "Текущий владелец", "Владелец")
#         cabinet = get_any(row, "Текущий кабинет (кабинет)", "Текущий кабинет", "Кабинет")

#         rows.append(
#             (
#                 inv,
#                 get(row, "Наименование оборудования"),
#                 get(row, "Дата принятия к учету"),
#                 owner,
#                 get(row, "Серийный номер"),
#                 get(row, "Дата постановки / последняя дата актуализации (опционально)"),
#                 cabinet,
#             )
#         )

#     conn.execute("DELETE FROM assets_reference")
#     conn.executemany(
#         """
#         INSERT OR REPLACE INTO assets_reference(inventory_number,name,accepted_date,owner,serial_number,last_update_date,cabinet)
#         VALUES(?,?,?,?,?,?,?)
#         """,
#         rows,
#     )
#     conn.commit()

#     print("IMPORTED ROWS:", len(rows))
#     return len(rows)


def _norm_location_to_cabinet(s: str) -> str:
    """
    Приводит 'Месторасположение' к формату, который понимает бот:
    - 'Кабинет 101' -> '101'
    - 'Серверная' -> 'СЕРВЕРНАЯ'
    - 'Конференц...' -> 'КОНФЕРЕНЦ-ЗАЛ'
    - 'Актовый...' -> 'АКТОВЫЙ ЗАЛ'
    """
    if s is None:
        return ""
    t = str(s).strip()
    low = t.lower()

    if "сервер" in low:
        return "СЕРВЕРНАЯ"
    if "конференц" in low:
        return "КОНФЕРЕНЦ-ЗАЛ"
    if "актов" in low:
        return "АКТОВЫЙ ЗАЛ"

    # Вытащить номер кабинета из строки
    m = re.search(r"(\d{1,4})", t)
    if m:
        return m.group(1)

    # если вдруг что-то нестандартное — оставим как есть
    return t


def import_reference_xlsx(conn: sqlite3.Connection, path: str) -> int:
    wb = load_workbook(path)
    # Ваш основной лист
    ws = wb["Лист_1"] if "Лист_1" in wb.sheetnames else wb.active

    print("IMPORT FILE:", path)
    print("SHEET:", ws.title, "max_row:", ws.max_row, "max_col:", ws.max_column)

    # Ищем шапку (в вашем файле она в первой строке)
    header_row_idx = 1
    header_values = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
    idx = {_norm_header(v): i for i, v in enumerate(header_values) if _norm_header(v)}

    def get(row, col_name: str) -> str:
        i = idx.get(_norm_header(col_name))
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else str(v).strip()

    def get_any(row, *col_names: str) -> str:
        for name in col_names:
            v = get(row, name)
            if v:
                return v
        return ""

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # В вашем файле: "Инв. №"
        inv_raw = get_any(row, "Инв. №", "Инв №", "Инвентарный номер", "Инвентарный №")
        inv = norm_inv(inv_raw)
        if not inv:
            continue

        name = get_any(row, "Основное средство", "Наименование оборудования", "Наименование")
        location = get_any(row, "Месторасположение", "Кабинет", "Локация")

        cabinet = _norm_location_to_cabinet(location)

        # В вашем справочнике сейчас нет владельца/серийника/дат — оставляем пустыми
        owner = get_any(
            row,
            "Сотрудник",                      # <-- ВАЖНО: ваша новая колонка
            "Текущий владелец (сотрудник)",
            "Текущий владелец",
            "Владелец",
            "Ответственный",
            "ФИО",
        )

        rows.append((
            inv,
            name,
            "",      # accepted_date
            owner,   # owner (если добавите колонку)
            "",      # serial_number
            "",      # last_update_date
            cabinet  # cabinet (нормализован)
        ))

    conn.execute("DELETE FROM assets_reference")
    conn.executemany("""
        INSERT OR REPLACE INTO assets_reference(
            inventory_number,name,accepted_date,owner,serial_number,last_update_date,cabinet
        )
        VALUES(?,?,?,?,?,?,?)
    """, rows)
    conn.commit()

    print("IMPORTED ROWS:", len(rows))
    return len(rows)


def rebuild_current_from_reference_and_history(conn: sqlite3.Connection) -> None:
    # 1) reference -> current
    conn.execute("DELETE FROM assets_current")

    now = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        """
        INSERT INTO assets_current(
            inventory_number,
            name,
            accepted_date,
            owner,
            serial_number,
            last_update_date,
            cabinet,
            updated_at
        )
        SELECT
            inventory_number,
            name,
            accepted_date,
            owner,
            serial_number,
            last_update_date,
            cabinet,
            ?
        FROM assets_reference
        """,
        (now,),
    )
    conn.commit()

    # 2) накатываем историю перемещений по порядку id
    cur = conn.execute("SELECT * FROM movements ORDER BY id ASC")
    for mv in cur.fetchall():
        inv = mv["inventory_number"]
        to_owner = mv["to_owner"]
        to_cab = mv["to_cabinet"]
        now2 = datetime.now().isoformat(timespec="seconds")

        conn.execute(
            """
            UPDATE assets_current
            SET owner = COALESCE(?, owner),
                cabinet = COALESCE(?, cabinet),
                updated_at = ?
            WHERE inventory_number = ?
            """,
            (to_owner, to_cab, now2, inv),
        )

    conn.commit()


def export_result_xlsx(conn: sqlite3.Connection, out_path: str) -> None:
    wb = Workbook()

    # --- Лист 1: os_reference ---
    ws1 = wb.active
    ws1.title = "os_reference"
    ws1.append(HEADERS)

    cur = conn.execute("""
        SELECT name, accepted_date, inventory_number, owner, serial_number, last_update_date, cabinet
        FROM assets_current
        ORDER BY inventory_number
    """)
    for row in cur.fetchall():
        ws1.append([
            row["name"],
            row["accepted_date"],
            row["inventory_number"],
            row["owner"],
            row["serial_number"],
            row["last_update_date"],
            row["cabinet"],
        ])

    # --- Лист 2: os_history ---
    ws2 = wb.create_sheet("os_history")
    ws2.append([
        "ID",
        "Дата перемещения",
        "Инициатор (имя)",
        "Инициатор (username)",
        "Инициатор (tg_id)",
        "Инвентарный номер",
        "Наименование оборудования",
        "Было: владелец",
        "Было: кабинет",
        "Стало: владелец",
        "Стало: кабинет",
        "Комментарий",
    ])

    cur2 = conn.execute("""
        SELECT
            m.id,
            m.moved_at,
            m.initiator_name,
            m.initiator_username,
            m.initiator_tg_id,
            m.inventory_number,
            c.name AS asset_name,
            m.from_owner,
            m.from_cabinet,
            m.to_owner,
            m.to_cabinet,
            m.comment
        FROM movements m
        LEFT JOIN assets_current c ON c.inventory_number = m.inventory_number
        ORDER BY m.id ASC
    """)

    for r in cur2.fetchall():
        # на случай старых записей, где initiator_name ещё пустой
        initiator_name = r["initiator_name"] if "initiator_name" in r.keys() else ""

        ws2.append([
            r["id"],
            r["moved_at"],
            initiator_name,
            r["initiator_username"],
            r["initiator_tg_id"],
            r["inventory_number"],
            r["asset_name"],
            r["from_owner"],
            r["from_cabinet"],
            r["to_owner"],
            r["to_cabinet"],
            r["comment"],
        ])

    wb.save(out_path)